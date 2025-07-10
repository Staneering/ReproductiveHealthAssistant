from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain.document_loaders import PyPDFLoader, TextLoader
from langchain.document_loaders.web_base import WebBaseLoader
from langchain.document_loaders import Docx2txtLoader
from langchain.vectorstores import Chroma
from langchain.chains.qa_with_sources.retrieval import RetrievalQAWithSourcesChain
from langchain.prompts import PromptTemplate
import chainlit as cl
from langchain_groq import ChatGroq
from sentence_transformers import SentenceTransformer
from langchain.embeddings.base import Embeddings
from typing import List
import os
import requests
from bs4 import BeautifulSoup
import tempfile
import openpyxl


"""class SentenceTransformerEmbeddings(Embeddings):
    def __init__(self, model_name: str):
        self.model = SentenceTransformer(model_name)

    def __call__(self, input: List[str]) -> List[List[float]]:
        # input is a list of strings
        input = [str(doc) for doc in input]
        return self.model.encode(input).tolist()

    #def embed_documents(self, documents: List[str]) -> List[List[float]]:
    #    documents = [str(doc) for doc in documents]
    #    return self.model.encode(documents).tolist()

    def embed_query(self, query: str) -> List[float]:
        if not isinstance(query, str):
            query = str(query)
        return self.model.encode([query])[0]"""
    
class SentenceTransformerEmbeddings:
    def __init__(self, model_name: str):
        self.model = SentenceTransformer(model_name)

    def __call__(self, input: List[str]) -> List[List[float]]:
        input = [str(doc) for doc in input]
        return self.model.encode(input).tolist()

        # Compatibility for older LangChain code
    def embed_documents(self, documents: List[str]) -> List[List[float]]:
        return self.__call__(documents)
    
    def embed_query(self, query: str) -> List[float]:
        return self.__call__([query])[0]


# Configuration
embedding = SentenceTransformerEmbeddings("sentence-transformers/all-MiniLM-L6-v2")
groq_api_key = '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'  # Add your Groq API key here
text_splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=100)
llm = ChatGroq(temperature=0, groq_api_key=groq_api_key, model_name="llama-3.3-70b-versatile")

# Drug analyzer system prompt
DRUG_ANALYZER_SYSTEM_PROMPT = """You are a knowledgeable and supportive sexual and reproductive health assistant. Your role is to provide accurate, respectful, and evidence-based information about sexual health, reproductive health, contraception, pregnancy, sexually transmitted infections (STIs), and related topics.

Your capabilities include:
- Explaining sexual and reproductive health concepts in clear, inclusive, and non-judgmental language
- Providing guidance on contraception, safe sex practices, and STI prevention
- Offering information about menstrual health, pregnancy, fertility, and reproductive anatomy
- Advising when to seek professional medical care or counseling
- Addressing common myths and misconceptions with factual information
- Supporting users of all genders, sexual orientations, and backgrounds

When responding:
1. Always prioritize user privacy, dignity, and safety
2. Use inclusive, respectful, and non-stigmatizing language
3. Clearly state when professional medical advice or services are needed
4. Reference trusted sources (such as WHO, CDC, NHS, or Planned Parenthood) when possible
5. Provide information suitable for a wide range of ages and backgrounds
6. Remind users that your guidance is for informational purposes only and does not replace professional healthcare

Remember: Your responses are for educational and support purposes only. Encourage users to consult qualified healthcare professionals for personalized advice, diagnosis, or treatment.
"""

# Predefined sources for drug analysis
DRUG_SOURCES = {
    # 'pdfs': [
    #     # Add paths to your PDF documents here
    #     # 'path/to/drug_reference.pdf',
    #     # 'path/to/pharmacology_textbook.pdf',
    # ],
    # 'docx': [
    #     # Add paths to your Word documents here
    #     # 'path/to/drug_guidelines.docx',
    #     # 'path/to/clinical_protocols.docx',
    # ],
    'web_urls': [
        # Add reliable drug information websites
        
        # Add more specific URLs as needed
        'https://www.drugs.com/aspirin.html'
    ]
}
def inject_urls(pth):
    dataframe = openpyxl.load_workbook(pth)
    dataframe1 = dataframe.active

    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            cell_value = col[row].value
            if cell_value and isinstance(cell_value, str):
                DRUG_SOURCES['web_urls'].append(cell_value.lstrip('m'))

inject_urls('C:\\Users\\Ibrah\\Documents\\fastapiProject\\chatbot\\sexualhealth.xlsx')
print("Loaded URLs:", DRUG_SOURCES['web_urls'])

def load_pdf_documents(pdf_paths: List[str]) -> List:
    """Load and process PDF documents"""
    documents = []
    for pdf_path in pdf_paths:
        if os.path.exists(pdf_path):
            try:
                loader = PyPDFLoader(pdf_path)
                docs = loader.load()
                for i, doc in enumerate(docs):
                    doc.metadata['source'] = f"{os.path.basename(pdf_path)}_page_{i+1}"
                    doc.metadata['source_type'] = 'pdf'
                documents.extend(docs)
            except Exception as e:
                print(f"Error loading PDF {pdf_path}: {e}")
    return documents


def load_docx_documents(docx_paths: List[str]) -> List:
    """Load and process Word documents"""
    documents = []
    for docx_path in docx_paths:
        if os.path.exists(docx_path):
            try:
                loader = Docx2txtLoader(docx_path)
                docs = loader.load()
                for i, doc in enumerate(docs):
                    doc.metadata['source'] = f"{os.path.basename(docx_path)}_section_{i+1}"
                    doc.metadata['source_type'] = 'docx'
                documents.extend(docs)
            except Exception as e:
                print(f"Error loading DOCX {docx_path}: {e}")
    return documents


def load_web_documents(urls: List[str]) -> List:
    """Load and process web documents"""
    documents = []
    for url in urls:
        try:
            # Use WebBaseLoader for better web scraping
            loader = WebBaseLoader(url)
            docs = loader.load()
            for i, doc in enumerate(docs):
                doc.metadata['source'] = f"{url}_section_{i+1}"
                doc.metadata['source_type'] = 'web'
                doc.metadata['url'] = url
            documents.extend(docs)
        except Exception as e:
            print(f"Error loading web content from {url}: {e}")
    return documents


def load_all_sources() -> List:
    """Load all predefined sources"""
    all_documents = []
    
    # Load PDFs
    # if DRUG_SOURCES['pdfs']:
    #     pdf_docs = load_pdf_documents(DRUG_SOURCES['pdfs'])
    #     all_documents.extend(pdf_docs)
    #     print(f"Loaded {len(pdf_docs)} PDF documents")
    
    # Load Word documents
    # if DRUG_SOURCES['docx']:
    #     docx_docs = load_docx_documents(DRUG_SOURCES['docx'])
    #     all_documents.extend(docx_docs)
    #     print(f"Loaded {len(docx_docs)} Word documents")
    
    # Load web sources
    if DRUG_SOURCES['web_urls']:
        web_docs = load_web_documents(DRUG_SOURCES['web_urls'])
        all_documents.extend(web_docs)
        print(f"Loaded {len(web_docs)} web documents")
    
    return all_documents


def create_vector_store(documents: List) -> Chroma:
    """Create vector store from documents"""
    if not documents:
        print("No documents to process")
        return None
    
    # Split documents
    splits = text_splitter.split_documents(documents)
    
    # Ensure each split has proper metadata
    for i, split in enumerate(splits):
        if 'source' not in split.metadata:
            split.metadata['source'] = f'document_chunk_{i}'
    
    # Create vector store
    texts = [doc.page_content for doc in splits]
    metadatas = [doc.metadata for doc in splits]
    
    vec_search = Chroma.from_texts(
        texts=texts,
        embedding=embedding,
        metadatas=metadatas,
        collection_name="drug_analyzer_store"
    )
    
    return vec_search, splits


def create_custom_prompt() -> PromptTemplate:
    """Create custom prompt template with system instructions"""
    template = f"""
{DRUG_ANALYZER_SYSTEM_PROMPT}

Context from relevant sources:
{{summaries}}

Human Question: {{question}}

Based on the provided context and your expertise in sexual and reproductive health, provide a comprehensive answer to the question.

Your response should:
1. Directly address the question with accurate, evidence-based information
2. Include relevant details about sexual health, reproductive health, contraception, pregnancy, STIs, or related topics as appropriate
3. Cite specific sources used in your explanation
4. Provide appropriate health disclaimers when necessary

Answer: Let me assist you with your sexual and reproductive health question. Please provide the context and your question, and I will do my best to help you.
"""
    
    return PromptTemplate(
        template=template,
        input_variables=["summaries", "question"]
    )


@cl.on_chat_start
async def start():
    """Initialize the sexual and reproductive health chatbot"""
    await cl.Message(content="🔬 **Sexual and Reproductive Health Chatbot Initializing...**\n\nLoading health databases and medical references...").send()

    try:
        # Load all predefined sources
        documents = load_all_sources()
        
        if not documents:
            await cl.Message(content="⚠️ **Warning**: No source documents found. Please add document paths to sexual and reproductive health sources in the code.").send()
            # Create empty vector store for demo purposes
            cl.user_session.set("docs", [])
            cl.user_session.set("vec_search", None)
            cl.user_session.set("chain", None)
            return
        
        # Create vector store
        vec_search, splits = create_vector_store(documents)
        
        # Store in session
        cl.user_session.set("docs", splits)
        cl.user_session.set("vec_search", vec_search)
        
        # Create custom QA chain with drug analyzer prompt
        custom_prompt = create_custom_prompt()
        
        chain = RetrievalQAWithSourcesChain.from_chain_type(
            llm=llm,
            chain_type="stuff",
            retriever=vec_search.as_retriever(search_kwargs={"k": 5}),
            return_source_documents=True,
            chain_type_kwargs={"prompt": custom_prompt}
        )
        
        cl.user_session.set("chain", chain)
        
        await cl.Message(
            content=f"""✅ **Drug Analyzer Ready!**

📊 **Loaded Sources:**
- Documents processed: {len(splits)} chunks
- Knowledge base initialized successfully

🧬 **Capabilities:**
- Information on contraception and safe sex practices
- Guidance on menstrual and reproductive health
- Answers about pregnancy, fertility, and sexual wellness
- Facts about sexually transmitted infections (STIs) and prevention
- Support for sexual health concerns and common questions

💬 **How to use:**
Ask me anything about sexual and reproductive health—such as contraception, menstrual cycles, pregnancy, STIs, sexual wellness, or related topics. I’ll provide accurate, respectful, and evidence-based answers with references to trusted sources.

⚠️ **Health Disclaimer:** This system is for educational and informational purposes only. Always consult qualified healthcare professionals for personal medical
"""
        ).send()
        
    except Exception as e:
        await cl.Message(content=f"❌ **Error initializing system**: {str(e)}").send()


@cl.on_message
async def main(message):
    """Handle user queries about drugs"""
    print('handlerr was triggered')
    chain = cl.user_session.get('chain')
    
    if not chain:
        await cl.Message(content="⚠️ System not properly initialized. Please restart the session.").send()
        return
    
    # Create callback handler for streaming
    cb = cl.AsyncLangchainCallbackHandler(
        stream_final_answer=True,
        answer_prefix_tokens=["FINAL", "ANSWER"]
    )
    cb.answer_reached = True
    
    try:
        # Process the query
        res = await chain.acall(message.content, callbacks=[cb])
        answer = res['answer']
        sources = res.get('sources', '').strip()
        
        # Process source elements
        source_elements = []
        docs = cl.user_session.get("docs", [])
        
        if sources and docs:
            docs_metadata = [doc.metadata for doc in docs]
            all_sources = [m.get('source', '') for m in docs_metadata]
            
            found_sources = []
            for source in sources.split(','):
                source_name = source.strip().replace('.', '')
                try:
                    index = all_sources.index(source_name)
                    text = docs[index].page_content
                    metadata = docs[index].metadata
                    
                    # Create rich source element with metadata
                    source_display = f"**Source:** {source_name}\n"
                    if 'source_type' in metadata:
                        source_display += f"**Type:** {metadata['source_type'].upper()}\n"
                    if 'url' in metadata:
                        source_display += f"**URL:** {metadata['url']}\n"
                    source_display += f"**Content:** {text[:500]}..."
                    
                    found_sources.append(source_name)
                    source_elements.append(cl.Text(content=source_display, name=source_name))
                except (ValueError, IndexError):
                    continue
            
            # Add sources summary to answer
            if found_sources:
                answer += f'\n\n📚 **Sources Referenced:** {", ".join(found_sources)}'
            else:
                answer += '\n\n📚 **Sources:** Analysis based on general sexual and reproductive health knowledge'
        
        # Update the streamed response or send new message
        if cb.has_streamed_final_answer:
            cb.final_stream.elements = source_elements
            await cb.final_stream.update()
        else:
            await cl.Message(content=answer, elements=source_elements).send()
            
    except Exception as e:
        await cl.Message(content=f"❌ **Error processing query**: {str(e)}").send()


if __name__ == "__main__":
    # You can add source documents here before running
    print("Sexual and Reproductive Health Chatbot")
    print("Add your source documents to DRUG_SOURCES configuration")
    print("Supported formats: PDF, DOCX, Web URLs")